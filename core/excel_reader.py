import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont   # ⭐ TextBlock 전용 폰트
from core.config import YELLOW_HEX


# ==============================
# 🔍 1) 원본에서 노란색 행 판별
# ==============================
def find_changed_rows(excel_path, sheet_name):
    wb = load_workbook(excel_path, data_only=True)

    if sheet_name not in wb.sheetnames:
        print(f"⚠️ '{sheet_name}' 시트를 찾을 수 없습니다.")
        return set()

    ws = wb[sheet_name]
    changed_rows = set()

    for row in ws.iter_rows(min_row=3):
        for cell in row:
            fill = cell.fill
            if fill and fill.start_color and fill.start_color.rgb:
                rgb = fill.start_color.rgb.upper()
                if rgb.endswith("FFFF00"):
                    changed_rows.add(cell.row)
                    break

    return changed_rows


# ==============================
# ⭐ 2) `_spot` 포함 텍스트만 색상 강조 (InlineFont 사용)
# ==============================
def apply_spot_richtext(cell):
    value = str(cell.value)
    if "_spot" not in value:
        return

    parts = [p.strip() for p in value.split(",")]

    rich = CellRichText()

    for i, part in enumerate(parts):

        if "_spot" in part:
            # ⭐ TextBlock 은 InlineFont 를 사용해야 함
            rich.append(TextBlock(
                text=part,
                font=InlineFont(color="D34E4E", b=True)  # 진하게 + 강조
            ))
        else:
            rich.append(TextBlock(
                text=part,
                font=InlineFont(color="000000")
            ))

        if i < len(parts) - 1:
            rich.append(TextBlock(text=", ", font=InlineFont(color="000000")))

    cell.value = rich


# ==============================
# 🟡 3) 로그 파일 엑셀 저장 + 하이라이트
# ==============================
def save_excel_with_highlight(df, path, changed_rows):
    df.to_excel(path, index=False, engine='openpyxl')

    wb = load_workbook(path)
    ws = wb.active

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # 🔧 자동 너비 조정
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter

        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[col_letter].width = (max_len + 2) * 1.2

    # 🟨 변경 행 전체 강조
    for src_row in changed_rows:
        log_row = src_row - 1
        if log_row >= 2:
            for cell in ws[log_row]:
                cell.fill = yellow_fill

    # 🎯 E열 `_spot` 하이라이트
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
        cell = row[0]
        if cell.value:
            apply_spot_richtext(cell)

    wb.save(path)
