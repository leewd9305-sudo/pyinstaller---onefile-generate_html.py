import os
import sys
import pandas as pd
import re
from datetime import datetime
import shutil
from tkinter import Tk, filedialog, messagebox
import unicodedata
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ==============================
# 🧭 PyInstaller 경로 인식
# ==============================
def resource_path(relative_path):
    if hasattr(sys, "_MEIPASS"):
        base_path = sys._MEIPASS
    else:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)


# ==============================
# 📂 엑셀 파일 선택
# ==============================
def select_excel_file():
    root = Tk()
    root.withdraw()

    file_path = filedialog.askopenfilename(
        title="제작가이드 엑셀 파일 선택",
        filetypes=[("Excel Files", "*.xlsx *.xls")]
    )

    if not file_path:
        raise FileNotFoundError("엑셀 파일이 선택되지 않았습니다!")

    print(f"\n📌 선택된 파일: {file_path}")
    return file_path


# ==============================
# 🔒 파일명 정리
# ==============================
def sanitize_filename(name: str) -> str:
    name = re.sub(r'[<>:"/\\|?*]', "_", str(name))
    return name.strip()


# ==============================
# 📌 원본 엑셀에서 노란색 셀이 포함된 row 찾기
# ==============================
YELLOW_HEX = ["FFFFFF00", "FFFF00"]

def find_changed_rows(excel_path, sheet_name):
    wb = load_workbook(excel_path, data_only=True)
    
    # 시트가 없는 경우 예외 처리
    if sheet_name not in wb.sheetnames:
        print(f"⚠️ 경고: '{sheet_name}' 시트를 엑셀 파일에서 찾을 수 없습니다.")
        return set()
        
    ws = wb[sheet_name]

    changed_rows = set()

    # 데이터가 시작하는 3번째 행(인덱스 기준 3)부터 반복
    for row in ws.iter_rows(min_row=3):  
        for cell in row:
            fill = cell.fill
            if fill and fill.start_color and fill.start_color.rgb:
                rgb = fill.start_color.rgb.upper()
                # 'AARRGGBB' 또는 'RRGGBB' 형태의 노란색 확인
                if rgb.endswith("FFFF00"): 
                    changed_rows.add(cell.row)
                    break

    return changed_rows


# ==============================
# 📘 엑셀 자동 셀 너비 조정 + 변경 row 노란색 강조
# ==============================
def save_excel_with_highlight(df, path, changed_rows):
    # changed_rows는 원본 엑셀 기준 행 번호 (min_row=3부터 시작)
    df.to_excel(path, index=False, engine='openpyxl')

    wb = load_workbook(path)
    ws = wb.active

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # 셀 너비 자동 조정 (기존 로직 유지)
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter

        for cell in col:
            try:
                # 한글(유니코드) 너비를 고려하여 조정 (기존 로직 유지)
                cell_length = len(str(cell.value))
                max_length = max(max_length, cell_length)
            except:
                pass

        ws.column_dimensions[col_letter].width = (max_length + 2) * 1.2

    # row 강조: 로그 파일의 헤더는 1행, 데이터는 2행부터 시작. 
    # 원본 row 번호 (>=3) - 1 (헤더) -> 로그 데이터의 행 번호
    for src_row in changed_rows:
        # 원본 엑셀의 3번째 행(데이터 시작)은 로그의 2번째 행이 됨
        log_row = src_row - 1  
        if log_row >= 2:
            for cell in ws[log_row]:
                cell.fill = yellow_fill

    wb.save(path)


# ==============================
# 📄 시트 → TXT(HTML) 변환
# ==============================
def generate_html_for_sheet(excel_file_path: str, sheet_name: str, output_dir: str, log_records: list):

    BLOB_BASE_URL = "https://huskb2bstorage.blob.core.windows.net/shopicus/dev_1/guide/03_make/page"
    TOOLTIP_BASE_URL = "https://huskb2bstorage.blob.core.windows.net/shopicus/dev_1/guide/test"

    print(f"\n🚀 [{sheet_name}] 변환 시작")

    # 폴더명 생성: '파일명 리스트(단색)' -> '단색'으로 변경
    folder_name_raw = sheet_name.replace("☆", "").strip()
    
    # 핵심 키워드 추출
    if "단색" in folder_name_raw:
        folder_name = "단색"
        tooltip_filename = "단색_툴팁.png"
        tooltip_alt = "단색 제작가이드"
        border_color = "#4DA3FF" # 블루
    elif "별색" in folder_name_raw:
        folder_name = "별색"
        tooltip_filename = "별색_툴팁.png"
        tooltip_alt = "별색 제작가이드"
        border_color = "#24CF7F" # 그린
    else:
        folder_name = "일반"
        tooltip_filename = "일반_툴팁.png" # 일반 시트용 툴팁 파일명 가정
        tooltip_alt = "일반 제작가이드"
        border_color = "#FFC107" # 옐로우 (임의 지정)


    sheet_output_dir = os.path.join(output_dir, folder_name)
    os.makedirs(sheet_output_dir, exist_ok=True)

    try:
        # header=None: 헤더 없이 데이터 로드
        df = pd.read_excel(excel_file_path, sheet_name=sheet_name, header=None, dtype=str)
        df = df.fillna("")
    except Exception as e:
        print(f"⚠️ 시트 '{sheet_name}' 로드 실패: {e}")
        return

    # 3번째 행(인덱스 2, 엑셀 3행)부터 데이터로 사용
    df = df.iloc[2:].copy()
    
    # 순번 열 (B열, 인덱스 1)이 비어있지 않은 행만 필터링
    df = df[df[1] != ""].copy()

    for _, row in df.iterrows():

        # 순번: B열 (인덱스 1)
        seq_raw = str(row[1]).strip()
        if not seq_raw:
            continue

        try:
            int(seq_raw)
            seq_str = seq_raw.zfill(2)
        except ValueError:
            seq_str = seq_raw

        # 제품명: C열 (인덱스 2)
        product_name = str(row[2]).strip()

        # 이미지 파일: D열(인덱스 3)부터 마지막 컬럼까지
        image_files = []
        for i in range(3, len(row)):
            val = row[i]
            if not val:
                continue

            clean_val = unicodedata.normalize("NFKC", str(val)).strip()
            # 파일명 유효성 검사 (필수 아님, 기존 로직 유지)
            image_files.append(clean_val)

        if not product_name or not image_files:
            continue

        safe_name = sanitize_filename(product_name)
        output_path = os.path.join(sheet_output_dir, f"{seq_str}_{safe_name}.txt")

        # HTML 구조 생성 (툴팁 및 색상 변수 사용)
        html = f"""
        <div style="width:100%; max-width:720px; margin:0 auto; padding:0 16px;
        display:flex; flex-direction:column; align-items:center; gap:20px;">

            <div style="border:4px solid {border_color}; border-radius:12px; width:100%;
                display:flex; flex-direction:column; align-items:center;
                padding-bottom:30px; position:relative;">

                <img src="{TOOLTIP_BASE_URL}/{tooltip_filename}"
                    alt="{tooltip_alt}"
                    style="position:absolute; top:15px; left:50%; transform:translateX(-50%);
                    width:130px; height:auto; z-index:10;">

                <h2 style="margin-top:75px; margin-bottom:30px;
                    font-size:20px; font-weight:600;">{product_name}</h2>
        """

        for i, file_name in enumerate(image_files, start=1):
            html += f"""
                <div style="margin-top:30px;">
                    <img src="{BLOB_BASE_URL}/{file_name}?ver={i}"
                        style="width:100%; max-width:450px;"
                        class="e-rte-image e-imginline">
                </div>
            """

        html += """
            </div>
        </div>
        """

        with open(output_path, "w", encoding="utf-8") as f:
            f.write(html)

        log_records.append({
            "시트명": sheet_name,
            "순번": seq_str,
            "제품명": product_name,
            "이미지_개수": len(image_files),
            "이미지_파일목록": ", ".join(image_files),
            "HTML_파일경로": output_path
        })

        print(f"✅ [{seq_str}] {product_name} → {output_path}")


# ==============================
# 🔍 단색 TXT → 콘텐츠 추출
# ==============================
def _extract_mono_content(html_path: str):
    # (기존 로직 유지)
    with open(html_path, "r", encoding="utf-8") as f:
        content = f.read()

    # <h2...> </h2> 태그 내부의 제품명 추출
    product_match = re.search(r'<h2[^>]*>(.*?)</h2>', content)
    product_name = product_match.group(1).strip() if product_match else ""

    # </h2> 태그 이후부터 마지막 </div> 앞까지의 이미지 콘텐츠 추출
    image_match = re.search(r'</h2[^>]*>([\s\S]*?)</div>\s*</div>\s*$', content)
    image_content = image_match.group(1).strip() if image_match else ""

    return product_name, image_content


# ==============================
# 🧱 단색/별색 공통 블록 생성
# ==============================
def _build_combined_block(product_name, image_content, tooltip_filename, tooltip_alt, border_color):
    # (기존 로직 유지)
    TOOLTIP_BASE_URL = "https://huskb2bstorage.blob.core.windows.net/shopicus/dev_1/guide/test"

    return f"""
    <div style="flex:1; text-align:center; display:flex;
        flex-direction:column; align-items:center;">

        <div style="border:4px solid {border_color}; border-radius:12px;
            width:100%; padding-bottom:30px; position:relative;">

            <img src="{TOOLTIP_BASE_URL}/{tooltip_filename}"
                alt="{tooltip_alt}"
                style="position:absolute; top:15px; left:50%; transform:translateX(-50%);
                width:130px; height:auto; z-index:10;">

            <h2 style="margin-top:75px; margin-bottom:30px;
                font-size:20px; font-weight:600;">{product_name}</h2>

            {image_content}

        </div>
    </div>
    """


# ==============================
# 🔗 단색 + 별색 병합 페이지 생성 (수정됨)
# ==============================
def generate_combined_html(output_dir):
    # 폴더명을 '파일명 리스트(단색)'에서 '단색'으로 변경
    mono_dir = os.path.join(output_dir, "단색")
    spot_dir = os.path.join(output_dir, "별색")
    combined_dir = os.path.join(output_dir, "combined")
    os.makedirs(combined_dir, exist_ok=True)

    if not os.path.exists(mono_dir) or not os.path.exists(spot_dir):
        print("⚠️ 병합 불가 — '단색' 또는 '별색' 폴더 없음")
        return

    # 단색 파일 목록 (순번 기준 정렬)
    mono_files = sorted(
        [f for f in os.listdir(mono_dir) if f.endswith(".txt")],
        key=lambda x: x.split("_", 1)[0]
    )

    # 별색 파일을 제품명_파일이름으로 맵핑
    spot_files_map = {
        os.path.splitext(f)[0].split("_", 1)[1]: f
        for f in os.listdir(spot_dir)
        if f.endswith(".txt")
    }

    for mono_file in mono_files:

        try:
            # 단색 파일에서 순번과 제품명 추출
            seq, product = os.path.splitext(mono_file)[0].split("_", 1)
        except:
            continue

        # 해당 제품명에 해당하는 별색 파일이 있는지 확인
        if product not in spot_files_map:
            continue

        spot_file = spot_files_map[product]

        # 1. 단색 TXT 파일에서 내용 추출
        mono_path = os.path.join(mono_dir, mono_file)
        mono_product_name, mono_image_content = _extract_mono_content(mono_path)

        # 2. 별색 TXT 파일에서 내용 추출
        spot_path = os.path.join(spot_dir, spot_file)
        spot_product_name, spot_image_content = _extract_mono_content(spot_path)


        # 왼쪽(단색) 블록 생성
        left_block = _build_combined_block(
            mono_product_name, mono_image_content,
            "단색_툴팁.png", "단색 제작가이드", "#4DA3FF"
        )

        # 오른쪽(별색) 블록 생성: 별색의 콘텐츠(spot_image_content) 사용
        right_block = _build_combined_block(
            spot_product_name, spot_image_content, # <-- 별색 콘텐츠 사용
            "별색_툴팁.png", "별색 제작가이드", "#24CF7F"
        )

        final_html = f"""
        <div style="width:100%; max-width:1420px; margin:0 auto; padding:0 16px;
        display:flex; justify-content:space-between; gap:30px; position:relative;">

            {left_block}

            <div style="position:absolute; top:0; left:50%; transform:translateX(-50%);
                width:1px; height:100%; background:#dcdcdc;"></div>

            {right_block}
        </div>
        """

        output_path = os.path.join(combined_dir, f"{seq}_{sanitize_filename(product)}.txt")

        with open(output_path, "w", encoding="utf-8") as f:
            f.write(final_html)

        print(f"✨ 병합 완료 → {output_path}")

    print("🎉 병합 TXT 생성 완료!")


# ==============================
# 🏁 메인 실행부 (기존 로직 유지)
# ==============================
if __name__ == "__main__":

    OUTPUT_DIR = resource_path("output")
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    try:
        EXCEL_FILE = select_excel_file()
    except:
        sys.exit(1)

    LOG_TIMESTAMP = datetime.now().strftime('%Y%m%d_%H%M%S')

    # 🔥 변경된 row 기록용 딕셔너리
    changed_rows_map = {}

    try:
        log_records = []

        excel_sheets = pd.ExcelFile(EXCEL_FILE)
        all_sheets = excel_sheets.sheet_names

        # 시트 필터링: '단색', '별색', '일반' 키워드가 포함된 시트를 유효한 시트로 간주
        valid_sheets = [s.strip() for s in all_sheets if ("단색" in s or "별색" in s or "일반" in s)]
        
        if not valid_sheets:
            messagebox.showwarning("경고", "처리할 유효한 시트('단색', '별색', '일반' 키워드 포함)를 찾지 못했습니다.")
            sys.exit(1)

        print(f"✅ 감지된 시트: {valid_sheets}")

        for sheet in valid_sheets:
            # 원본 엑셀에서 변경 row 찾기
            changed_rows_map[sheet] = find_changed_rows(EXCEL_FILE, sheet)

            # TXT 파일 생성 및 로그 기록
            generate_html_for_sheet(EXCEL_FILE, sheet, OUTPUT_DIR, log_records)

        # 로그 생성 및 분리
        if log_records:
            log_df = pd.DataFrame(log_records)

            # 전체 로그 (참고용)
            LOG_XLSX_ALL = os.path.join(OUTPUT_DIR, f"html_log_all_{LOG_TIMESTAMP}.xlsx")
            log_df.to_excel(LOG_XLSX_ALL, index=False, engine='openpyxl') 

            # 단색, 별색, 일반 로그 분리 및 강조
            mono_df = log_df[log_df["시트명"].str.contains("단색", na=False)]
            spot_df = log_df[log_df["시트명"].str.contains("별색", na=False)]
            normal_df = log_df[log_df["시트명"].str.contains("일반", na=False)]

            
            # --- 단색 로그 저장 ---
            # '단색' 시트의 정확한 이름을 찾아 변경된 행 적용
            mono_sheet_name = next((s for s in valid_sheets if "단색" in s), None)
            if not mono_df.empty and mono_sheet_name:
                save_excel_with_highlight(
                    mono_df, 
                    os.path.join(OUTPUT_DIR, f"log_mono_{LOG_TIMESTAMP}.xlsx"),
                    changed_rows_map.get(mono_sheet_name, set())
                )

            # --- 별색 로그 저장 ---
            spot_sheet_name = next((s for s in valid_sheets if "별색" in s), None)
            if not spot_df.empty and spot_sheet_name:
                save_excel_with_highlight(
                    spot_df, 
                    os.path.join(OUTPUT_DIR, f"log_spot_{LOG_TIMESTAMP}.xlsx"),
                    changed_rows_map.get(spot_sheet_name, set())
                )
            
            # --- 일반 로그 저장 ---
            normal_sheet_name = next((s for s in valid_sheets if "일반" in s), None)
            if not normal_df.empty and normal_sheet_name:
                save_excel_with_highlight(
                    normal_df, 
                    os.path.join(OUTPUT_DIR, f"log_normal_{LOG_TIMESTAMP}.xlsx"),
                    changed_rows_map.get(normal_sheet_name, set())
                )
        
        # 단색 + 별색 병합 페이지 생성
        generate_combined_html(OUTPUT_DIR)

        # ZIP 압축 생성 및 완료 메시지
        downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
        zip_filename = f"husk_guide_output_{LOG_TIMESTAMP}.zip"
        zip_path_base = os.path.join(downloads_path, zip_filename).replace(".zip", "")

        shutil.make_archive(
            base_name=zip_path_base,
            format="zip",
            root_dir=OUTPUT_DIR
        )

        messagebox.showinfo(
            "완료",
            f"제작가이드 변환이 완료되었습니다!\n압축 파일 위치:\n{zip_path_base}.zip"
        )

    except Exception:
        import traceback
        error_path = os.path.join(os.path.dirname(__file__), "error_log.txt")
        with open(error_path, "w", encoding="utf-8") as f:
            f.write(traceback.format_exc())
        
        # Tkinter 오류 메시지 박스 추가
        messagebox.showerror(
            "오류 발생",
            f"스크립트 실행 중 오류가 발생했습니다. 자세한 내용은 다음 파일을 확인하세요:\n{error_path}"
        )
        print(f"⚠️ 오류 발생 → {error_path}")